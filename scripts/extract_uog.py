"""Однократное извлечение текстов/таблиц UOG-HSE-PR-007 для сверки с NOVA Safety."""
from __future__ import annotations

import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

ROOT = Path(r"D:\gp_2026_cad26\docs\UOG-PR-007")
OUT = ROOT / "_machine_read"
W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def docx_to_text(path: Path) -> str:
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml")
    root = ET.fromstring(xml)
    chunks: list[str] = []
    for t in root.iter(W + "t"):
        if t.text:
            chunks.append(t.text)
        if t.tail:
            chunks.append(t.tail)
    return "".join(chunks)


def safe_name(path: Path) -> str:
    s = path.stem
    s = re.sub(r"[^\w\-]+", "_", s, flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")[:80]
    return f"{s}_{path.suffix[1:]}"


def dump_xlsx(path: Path, dest_json: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    data: dict = {"file": path.name, "sheets": {}}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows: list[list[str | None]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 80:
                break
            rows.append([str(c) if c is not None else None for c in row])
        data["sheets"][sn] = rows
    dest_json.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    index: list[dict] = []

    for path in sorted(ROOT.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if "_machine_read" in path.parts or path.suffix.lower() not in {".docx", ".xlsx"}:
            continue
        if "Архив" in path.parts or "Arhiv" in path.parts:
            continue

        tag = safe_name(path)
        rel = str(path.relative_to(ROOT))

        if path.suffix.lower() == ".docx":
            try:
                text = docx_to_text(path)
                tpath = OUT / f"{tag}.txt"
                tpath.write_text(text, encoding="utf-8")
                index.append({"type": "docx", "source": rel, "out": str(tpath.relative_to(OUT))})
            except Exception as e:
                index.append({"type": "docx", "source": rel, "error": str(e)})

        elif path.suffix.lower() == ".xlsx":
            try:
                jpath = OUT / f"{tag}.json"
                dump_xlsx(path, jpath)
                index.append({"type": "xlsx", "source": rel, "out": str(jpath.relative_to(OUT))})
            except Exception as e:
                index.append({"type": "xlsx", "source": rel, "error": str(e)})

    (OUT / "index.json").write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")
    print("files:", len(index), "->", OUT)


if __name__ == "__main__":
    main()
